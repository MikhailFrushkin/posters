import urllib


def convert_to_yandex_disk_link(path: str) -> str:
    # Удаляем префикс "disk:/", если он присутствует
    if path.startswith("disk:/"):
        path = path[6:]

    # Добавляем пробел после "ANIKOYA"
    path = path.replace(" ANIKOYA ", " ANIKOYA%20")

    # Заменяем пробелы на "%20"
    path = path.replace(" ", "%20")

    # Добавляем префикс "https://disk.yandex.ru/client/disk/"
    link = "https://disk.yandex.ru/client/disk/" + path

    return link


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows


# Функция для обновления ссылки на Яндекс.Диск
def update_yandex_disk_link(row):
    path = row['Путь']
    url = convert_to_yandex_disk_link(path)
    return url


# Чтение эксель-файла
df = pd.read_excel('Пути к артикулам 1.xlsx')
df['URL'] = df.apply(update_yandex_disk_link, axis=1)

# Создание нового эксель-файла
workbook = Workbook()
sheet = workbook.active

for r in dataframe_to_rows(df, index=False, header=True):
    sheet.append(r)

# Добавление ссылок в колонку "URL"
for row in sheet.iter_rows(min_row=2, min_col=sheet.max_column, max_col=sheet.max_column):
    cell = row[0]
    hyperlink = cell.value
    cell.value = None
    cell.hyperlink = hyperlink
    cell.font = Font(underline="single", color="0563C1")

# Сохранение обновленного эксель-файла
workbook.save('обновленный_файл.xlsx')
